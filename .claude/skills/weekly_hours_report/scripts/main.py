#!/usr/bin/env python3
"""Weekly Hours + Monthly Billing — orchestrator.

One command produces ONE .xlsx with two tabs ('Weekly Hours' + 'Invoice'),
reading everything from the same "PTO Tracker Softtek" Google Sheet (via core.py).

Rates come from a per-collaborator rates.json (the Sheet has no rate column). If
any collaborator is missing a rate, a template is written and the run stops so
the user can fill it in.

    /usr/bin/python3 scripts/main.py --check-rates        # validate rates only
    /usr/bin/python3 scripts/main.py --month Jul --year 2026
"""
import argparse
import sys
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

from core import fetch_sheet, fetch_team_roster, read_pto, read_team, build_report, write_weekly_sheet, DEFAULT_SHEET_ID
from billing import build_billing, write_invoice_sheet, load_rates, check_rates, write_rates_template, RATES_PATH_DEFAULT, TEMPLATE_PATH

DRIVE_OUTPUTS_REL = "Shared drives/Meta - STK/Project Tracking/Automation/Automation Outputs"
DRIVE_OUTPUT_SUBFOLDER = "Weekly Hours Report"

def resolve_output_path(month, year, override):
    fname = f"Weekly-Hours-Billing-{month}-{year}.xlsx"
    if override:
        return Path(override)
    cloud = Path.home() / "Library/CloudStorage"
    if cloud.exists():
        for child in sorted(cloud.iterdir()):
            if not child.name.startswith("GoogleDrive-"):
                continue
            outputs = child / DRIVE_OUTPUTS_REL
            if outputs.exists():
                folder = outputs / DRIVE_OUTPUT_SUBFOLDER
                folder.mkdir(exist_ok=True)
                return folder / fname
    print("  WARN: Shared drive 'Automation Outputs' not found locally; saving to your home folder instead")
    return Path.home() / fname

def _validate_rates(team, rates_path):
    """Return True if every collaborator has a rate; else write a template, report, and return False."""
    rates = load_rates(rates_path)
    missing = check_rates(team, rates)
    if missing:
        write_rates_template(rates_path, team, rates)
        print(f"\nMISSING RATES ({len(missing)} of {len(team)} collaborators) — add their hourly rate:")
        for n in missing:
            print(f"  - {n}")
        print(f"\nEdit {rates_path} (a template with all collaborators was written/updated),")
        print("fill in the rates, then re-run.")
        return False, rates
    return True, rates

def main():
    ap = argparse.ArgumentParser(description="Generate Weekly Hours + Monthly Billing (one .xlsx, two tabs)")
    ap.add_argument("--month", help="Month name (Jan-Dec), default current")
    ap.add_argument("--year", type=int, help="Year, default current")
    ap.add_argument("--sheet-id", default=DEFAULT_SHEET_ID, help="Google Sheet ID (collaborators + PTO)")
    ap.add_argument("--sheet-xlsx", help="Use an already-fetched local Sheet xlsx instead of downloading")
    ap.add_argument("--rates", help="Path to rates.json (default: scripts/rates.json)")
    ap.add_argument("--check-rates", action="store_true", help="Only validate rates.json (no month needed) and exit")
    ap.add_argument("--output", help="Output Excel path (default: Shared drive 'Weekly Hours Report' folder)")
    args = ap.parse_args()

    rates_path = Path(args.rates) if args.rates else RATES_PATH_DEFAULT

    # --- rates-only check (month-independent) ---
    if args.check_rates:
        print("Validating rates.json against the collaborator roster...")
        team = fetch_team_roster(sheet_id=args.sheet_id)
        print(f"  {len(team)} collaborators in Team allocation 2026")
        ok, _ = _validate_rates(team, rates_path)
        if ok:
            print(f"\nAll {len(team)} collaborators have a rate. Ready to generate.")
        sys.exit(0 if ok else 2)

    now = datetime.now()
    month = args.month or ["Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec"][now.month-1]
    year = args.year or now.year

    print(f"Generating Weekly Hours + Billing for {month} {year}...")

    if args.sheet_xlsx:
        sheet_xlsx = Path(args.sheet_xlsx)
        print(f"Using local Sheet export: {sheet_xlsx}")
    else:
        print("Fetching Google Sheet (collaborators + PTO)...")
        sheet_xlsx = fetch_sheet(month, year, sheet_id=args.sheet_id)
        print(f"  Using local copy: {sheet_xlsx}")

    print("Reading PTO (monthly tab)...")
    pto = read_pto(month, year, pto_path=sheet_xlsx)
    print(f"  Found {len(pto)} employees in PTO")

    print("Reading Team Allocation tab...")
    team = read_team(team_path=sheet_xlsx)
    print(f"  Found {len(team)} collaborators")

    # --- rates must be complete before we bill ---
    ok, rates = _validate_rates(team, rates_path)
    if not ok:
        sys.exit(2)

    print("Building weekly hours...")
    weekly = build_report(pto, team, month, year)
    tw = len(weekly["q1"]) + len(weekly["q2"]) + len(weekly["q3"])
    print(f"  Q1:{len(weekly['q1'])} Q2:{len(weekly['q2'])} Q3:{len(weekly['q3'])} Total:{tw}")

    print("Building billing...")
    billing = build_billing(weekly, rates)
    print(f"  Subtotal: ${billing['subtotal']:,.2f}  Discount {billing['discountRate']*100:.0f}%: ${billing['discountAmount']:,.2f}  Total: ${billing['total']:,.2f}")

    print("Writing combined workbook (Weekly Hours + Invoice from template)...")
    if not TEMPLATE_PATH.exists():
        print(f"ERROR: invoice template not found at {TEMPLATE_PATH}")
        sys.exit(3)
    wb = load_workbook(TEMPLATE_PATH)              # the invoice comes pre-formatted from the template
    invoice_ws = wb[wb.sheetnames[0]]
    invoice_ws.title = "Invoice"
    work_refs = write_weekly_sheet(wb, weekly)    # add the Weekly Hours tab; returns Work Hrs cell refs
    wb.move_sheet("Weekly Hours", offset=-1)       # show Weekly Hours as the first tab
    write_invoice_sheet(invoice_ws, billing, month, year, work_refs=work_refs)

    out_path = resolve_output_path(month, year, args.output)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    print(f"\nDone! Output: {out_path}")

if __name__ == "__main__":
    main()
