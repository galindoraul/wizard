"""Monthly Billing — library used by main.py.

Turns each collaborator's worked hours (from the weekly report) into a
Softtek → Meta invoice. Reads the same data as the Weekly Hours report (via
core.py). Billing RATES come from a per-collaborator `rates.json` (the Google
Sheet has no rate column).
"""
import calendar
import copy
import json
from datetime import date
from pathlib import Path

from openpyxl.styles import Font, Alignment, Border, PatternFill

from core import normalize_value, get_month_number, get_month_name

RATES_PATH_DEFAULT = Path(__file__).parent.parent / "assets" / "rates.json"
DISCOUNT_RATE = 0.02  # 2% volume discount


# ===========================================================================
# Rates — per-collaborator, loaded from rates.json
# ===========================================================================
def load_rates(path):
    """Return {normalized short name -> rate(float) or None}. Missing file -> {}."""
    path = Path(path)
    if not path.exists():
        return {}
    try:
        raw = json.loads(path.read_text())
    except Exception:
        return {}
    out = {}
    for name, val in raw.items():
        try:
            out[normalize_value(name)] = float(val)
        except (TypeError, ValueError):
            out[normalize_value(name)] = None
    return out

def check_rates(team_data, rates):
    """Return the list of collaborator short names that have no positive rate."""
    missing = []
    for c in team_data:
        name = c["short_name"]
        if not name:
            continue
        r = rates.get(normalize_value(name))
        if not r or r <= 0:
            missing.append(name)
    return missing

def write_rates_template(path, team_data, existing):
    """Write/refresh rates.json with every collaborator (existing values kept, missing -> 0)."""
    data = {}
    for c in team_data:
        name = c["short_name"]
        if not name:
            continue
        r = existing.get(normalize_value(name))
        data[name] = r if (r and r > 0) else 0
    Path(path).write_text(json.dumps(data, indent=2, ensure_ascii=False))


# ===========================================================================
# Build billing from the weekly report + rates
# ===========================================================================
def _rate_for(emp, rates):
    """Principal is billed at their own rate; a backup at the rate of whoever they cover."""
    key = normalize_value(emp.get("coveringFor", "") if emp.get("isBackup")
                          else emp["fullName"].lstrip("↳ ").strip())
    return rates.get(key) or 0.0

def build_billing(weekly_report, rates):
    def build_section(employees, role_key):
        rows = []
        for emp in employees:
            if emp.get("isSeparator"):
                continue
            rate = _rate_for(emp, rates)
            if rate == 0:
                print(f"WARN no rate for {emp['fullName']} — billed at $0")
            qty = emp["workHrs"]
            rows.append({
                "fullName": emp["fullName"], "tag": emp["tag"], "role": emp["role"],
                "product": emp["product"], "pilar": emp["pilar"],
                "qty": qty, "unit": "Hrs", "description": emp["tag"],
                "rate": rate, "amount": qty * rate,
                "isBackup": emp.get("isBackup", False), "coveringFor": emp.get("coveringFor"),
                "isSeparator": False, "roleKey": role_key,
                "_emp": emp,  # original weekly emp — lets the invoice reference its Work Hrs cell
            })
        return rows

    q1 = build_section(weekly_report["q1"], "q1")
    q2 = build_section(weekly_report["q2"], "q2")
    q3 = build_section(weekly_report["q3"], "q3")
    subtotal = sum(e["amount"] for e in q1 + q2 + q3)
    discount_amount = subtotal * DISCOUNT_RATE
    total = subtotal - discount_amount
    return {"q1": q1, "q2": q2, "q3": q3, "subtotal": subtotal,
            "discountRate": DISCOUNT_RATE, "discountAmount": discount_amount, "total": total}


# ===========================================================================
# Amount-in-words (invoice needs the total spelled out)
# ===========================================================================
_ONES = ["zero","one","two","three","four","five","six","seven","eight","nine","ten",
         "eleven","twelve","thirteen","fourteen","fifteen","sixteen","seventeen","eighteen","nineteen"]
_TENS = ["","","twenty","thirty","forty","fifty","sixty","seventy","eighty","ninety"]

def _under_1000(n):
    if n < 20: return _ONES[n]
    if n < 100: return _TENS[n // 10] + ("-" + _ONES[n % 10] if n % 10 else "")
    return _ONES[n // 100] + " hundred" + (" " + _under_1000(n % 100) if n % 100 else "")

def _int_words(n):
    if n == 0: return "zero"
    parts = []
    for div, name in [(1_000_000, "million"), (1000, "thousand"), (1, "")]:
        if n >= div:
            parts.append(_under_1000(n // div) + ((" " + name) if name else ""))
            n %= div
    return " ".join(parts)

def amount_to_words(total):
    dollars = int(total)
    cents = int(round((total - dollars) * 100))
    return f"{_int_words(dollars).upper()} DOLLARS AND {(_int_words(cents).upper() if cents else 'ZERO')} CENTS"


# ===========================================================================
# Invoice worksheet — fill the "Monthly Billing Report" template
#
# Instead of drawing the invoice from scratch, we open the pre-formatted
# template (assets/Monthly Billing Report - Template.xlsx) and only fill the
# dynamic parts: the DATE, the QA Analyst I/II/III line items, and the totals
# formulas. Ported from the Google Apps Script `InvoiceDetailsWriter`: find each
# role's section marker, insert one row per employee under it, and write the row.
# ===========================================================================
TEMPLATE_PATH = Path(__file__).parent.parent / "assets" / "Monthly Billing Report - Template.xlsx"

# Template column layout (see the header row 17: QTY|UN|DESCRIPTION … RATE|AMOUNT)
QTY_COL = 1          # A
UNIT_COL = 2         # B
DESC_COL = 3         # C  (merged C:E on data rows)
DESC_COL_END = 5     # E
RATE_COL = 10        # J
AMOUNT_COL = 11      # K  (merged K:L on data rows)
AMOUNT_COL_END = 12  # L
MAX_COL = 12
SECTION_MARKER_COL = DESC_COL  # section labels ("QA Analyst I") live in column C
DATE_CELL = "I8"     # value cell next to the "DATE" label (merged I8:L8)
SERVICE_PERIOD_CELL = "C18"  # blank line under the table header (merged C18:F18)

INVOICE_FONT = "Times New Roman"  # everything we write into the invoice uses this typeface

def _tnr(cell, bold=None):
    """Retype a written cell in Times New Roman, keeping its size/color; `bold`
    overrides the weight when given (collaborators are written un-bolded)."""
    f = cell.font
    cell.font = Font(name=INVOICE_FONT, size=f.size, color=f.color,
                     bold=f.bold if bold is None else bold,
                     italic=f.italic, underline=f.underline)

# section label (col C) -> the weekly-report role bucket it bills
ROLE_SECTIONS = [
    ("QA Analyst III", "q3"),
    ("QA Analyst II", "q2"),
    ("QA Analyst I", "q1"),
]

def invoice_date(month, year):
    """Last calendar day of the billed month (shown as the invoice DATE)."""
    month_num = get_month_number(month)
    last_day = calendar.monthrange(year, month_num)[1]
    return date(year, month_num, last_day)

def _find_row_by_value(ws, value, col):
    """First row whose cell in `col` equals `value` exactly, or None."""
    for row in range(1, ws.max_row + 1):
        if ws.cell(row, col).value == value:
            return row
    return None

def _copy_cell_style(src, dst):
    dst.font = copy.copy(src.font)
    dst.fill = copy.copy(src.fill)
    dst.border = copy.copy(src.border)
    dst.alignment = copy.copy(src.alignment)
    dst.number_format = src.number_format

def _insert_styled_rows(ws, at_row, count, style_row):
    """Insert `count` rows at `at_row`, cloning styles + merges from `style_row`.

    openpyxl's insert_rows shifts existing cell values down but leaves merged
    ranges and formula references untouched and the new rows unstyled. So we move
    the merges at/below the insertion point down ourselves, then restyle the fresh
    rows from the template's pre-formatted placeholder row (`style_row`).
    """
    if count <= 0:
        return
    style_merges = [(m.min_col, m.max_col) for m in ws.merged_cells.ranges
                    if m.min_row == style_row and m.max_row == style_row]
    below = [m for m in list(ws.merged_cells.ranges) if m.min_row >= at_row]
    for m in below:
        ws.unmerge_cells(str(m))
    ws.insert_rows(at_row, count)
    for m in below:
        ws.merge_cells(start_row=m.min_row + count, end_row=m.max_row + count,
                       start_column=m.min_col, end_column=m.max_col)
    for i in range(count):
        r = at_row + i
        for col in range(1, MAX_COL + 1):
            _copy_cell_style(ws.cell(style_row, col), ws.cell(r, col))
        for c0, c1 in style_merges:
            ws.merge_cells(start_row=r, end_row=r, start_column=c0, end_column=c1)

def _clear_row(ws, row):
    """Turn `row` into a clean blank spacer — no merges, value, border or fill."""
    for m in [m for m in list(ws.merged_cells.ranges) if m.min_row == row and m.max_row == row]:
        ws.unmerge_cells(str(m))
    for col in range(1, MAX_COL + 1):
        c = ws.cell(row, col)
        c.value = None
        c.border = Border()
        c.fill = PatternFill()

def _write_employee_row(ws, row, emp, work_refs, tag_refs):
    """Fill one line item (QTY / UN / DESCRIPTION / RATE / AMOUNT) at `row`.

    Everything is typed in Times New Roman; collaborator names are NOT bolded.
    """
    ref = work_refs.get(id(emp.get("_emp")))
    # QTY: live link to the Weekly Hours "Work Hrs" cell when available, else static.
    qty = ws.cell(row, QTY_COL)
    qty.value = f"={ref}" if ref else emp["qty"]
    qty.alignment = Alignment(horizontal="right", vertical=qty.alignment.vertical)
    _tnr(qty)

    unit = ws.cell(row, UNIT_COL); unit.value = emp["unit"]; _tnr(unit)  # "Hrs"

    name = emp["fullName"].lstrip("↳ ").strip()
    # DESCRIPTION: live link to the Weekly Hours "Tag" cell when available, so an edit
    # to a collaborator's Tag flows into the invoice; else the static "Name - Product - Pilar".
    tag_ref = tag_refs.get(id(emp.get("_emp")))
    desc = ws.cell(row, DESC_COL)
    desc.value = f"={tag_ref}" if tag_ref else f"{name} - {emp['product']} - {emp['pilar']}"
    _tnr(desc, bold=False)  # collaborators never bold
    if emp["isBackup"]:
        desc.alignment = Alignment(indent=1, horizontal=desc.alignment.horizontal,
                                   vertical=desc.alignment.vertical)

    rate = ws.cell(row, RATE_COL)
    rate.value = emp["rate"]; rate.number_format = '"$"0.00'; _tnr(rate)

    # AMOUNT = QTY * RATE, so it tracks any manual edit to the linked hours.
    amt = ws.cell(row, AMOUNT_COL)
    amt.value = f"=A{row}*J{row}"; amt.number_format = '"$"#,##0.00'; _tnr(amt)

def _write_totals(ws, billing, amount_rows):
    """Rewrite the template's SUB-TOTAL / discount / TOTAL formulas for the final
    row positions (they moved as rows were inserted, so their refs need fixing)."""
    st = _find_row_by_value(ws, "SUB-TOTAL", 8)
    disc = _find_row_by_value(ws, "Volume Discount", 8)
    tot = _find_row_by_value(ws, "TOTAL", 8)

    sc = ws.cell(st, AMOUNT_COL_END)
    sc.value = f"=SUM(K{amount_rows[0]}:L{amount_rows[-1]})" if amount_rows else billing["subtotal"]
    sc.number_format = "#,##0.00"; _tnr(sc)

    if disc is not None:
        dc = ws.cell(disc, AMOUNT_COL_END)
        dc.value = f"=(L{st}*{DISCOUNT_RATE})"; dc.number_format = "#,##0.00"; _tnr(dc)
        words = ws.cell(disc, 1); words.value = amount_to_words(billing["total"]); _tnr(words)  # amount in words
    if tot is not None:
        tc = ws.cell(tot, AMOUNT_COL_END)
        tc.value = f"=L{st}-L{disc}"; tc.number_format = "#,##0.00"; _tnr(tc)

def write_invoice_sheet(ws, billing, month, year, work_refs=None, tag_refs=None):
    """Fill the Monthly Billing template's invoice sheet with the QA line items.

    ``ws`` is the template sheet (already loaded into the combined workbook and
    renamed to "Invoice"). Only the dynamic parts are written — the DATE, the
    QA Analyst I/II/III employee rows, and the totals formulas — everything else
    (bill-to / ship-to, PO number, footer, styling) comes straight from the
    template instead of being drawn from scratch.

    When ``work_refs`` (id(emp) -> "'Weekly Hours'!<cell>") is provided, each QTY
    cell becomes a live formula pointing at that collaborator's Work Hrs cell, and
    AMOUNT / SUB-TOTAL / discount / TOTAL are formulas too — so a manual edit in
    the Weekly Hours tab flows straight into the invoice without re-running.
    """
    work_refs = work_refs or {}
    tag_refs = tag_refs or {}
    month_num = get_month_number(month)
    last_day = calendar.monthrange(year, month_num)[1]

    # DATE — last calendar day of the billed month (centered across I8:L8).
    d = ws[DATE_CELL]
    d.value = invoice_date(month, year); d.number_format = "M/D/YYYY"
    d.alignment = Alignment(horizontal="center", vertical=d.alignment.vertical)
    _tnr(d)

    # Services Period line (e.g. "Services Period: 7/1/2026 to 7/31/2026").
    sp = ws[SERVICE_PERIOD_CELL]
    sp.value = f"Services Period: {month_num}/1/{year} to {month_num}/{last_day}/{year}"
    _tnr(sp)

    # Pass 1 — make room under each section marker: n employee rows + 1 blank
    # spacer row (so sections aren't cramped). The template ships 1 placeholder row,
    # so we insert n more. Bottom-up (III -> II -> I) so inserting under a lower
    # section never shifts a higher section marker we still have to locate.
    for label, role_key in ROLE_SECTIONS:
        marker_row = _find_row_by_value(ws, label, SECTION_MARKER_COL)
        if marker_row is None:
            raise ValueError(f"Section marker not found in template: {label}")
        _insert_styled_rows(ws, marker_row + 2, len(billing[role_key]), marker_row + 1)

    # Pass 2 — fill rows (layout is final now, so formula row refs are stable).
    amount_rows = []
    for label, role_key in ROLE_SECTIONS:
        marker_row = _find_row_by_value(ws, label, SECTION_MARKER_COL)
        employees = billing[role_key]
        for i, emp in enumerate(employees):
            row = marker_row + 1 + i
            _write_employee_row(ws, row, emp, work_refs, tag_refs)
            amount_rows.append(row)
        _clear_row(ws, marker_row + 1 + len(employees))  # blank spacer after the section
    amount_rows.sort()

    _write_totals(ws, billing, amount_rows)
    return ws
