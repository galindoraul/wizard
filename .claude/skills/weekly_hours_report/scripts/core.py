"""Weekly Hours Report — consolidated logic.

Single module holding everything the report needs, imported by main.py:
  - date/month helpers
  - fetch_sheet(): download the Google Sheet tabs into a local .xlsx (fresh, parallel)
  - read_pto() / read_team(): parse the local .xlsx
  - build_report(): combine PTO + team into Q1/Q2/Q3 structures
  - export_excel(): write the formatted, color-coded workbook

Everything (collaborators AND monthly PTO) comes from a single Google Sheet
("PTO Tracker Softtek"). Data is pulled via `meta google.sheets read`.
"""
import json
import re
import subprocess
import tempfile
import unicodedata
from calendar import monthrange
from collections import defaultdict
from concurrent.futures import ThreadPoolExecutor
from datetime import date
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Border, Side, Font, Alignment
from openpyxl.utils import get_column_letter


# ===========================================================================
# Date / month helpers
# ===========================================================================
MONTHS = {"Jan":1,"January":1,"Feb":2,"February":2,"Mar":3,"March":3,"Apr":4,"April":4,"May":5,"Jun":6,"June":6,"Jul":7,"July":7,"Aug":8,"August":8,"Sep":9,"Sept":9,"September":9,"Oct":10,"October":10,"Nov":11,"November":11,"Dec":12,"December":12}
MONTH_NAMES_SHORT = ["Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec"]

def normalize_value(value):
    if not value: return ""
    s=str(value); s=unicodedata.normalize("NFD",s); s="".join(c for c in s if unicodedata.category(c)!="Mn")
    return s.lower().replace(" ","").strip()

def get_month_number(month):
    n=MONTHS.get(month)
    if not n: raise ValueError(f"Invalid month: {month}")
    return n

def get_month_name(month_number, short=True):
    return MONTH_NAMES_SHORT[month_number-1] if short else ["January","February","March","April","May","June","July","August","September","October","November","December"][month_number-1]

def get_week_of_month(d):
    first=date(d.year,d.month,1); offset=first.weekday(); return (d.day+offset-1)//7+1

def build_month_context(year, month):
    month_num=month if isinstance(month,int) else get_month_number(month)
    month_name=get_month_name(month_num,True); days_in_month=monthrange(year,month_num)[1]; weeks_map={}
    for day in range(1,days_in_month+1):
        d=date(year,month_num,day)
        if d.weekday()>=5: continue
        wn=get_week_of_month(d); weeks_map.setdefault(wn,[]).append(day)
    weeks=[]
    for wn in sorted(weeks_map):
        workdays=weeks_map[wn]; start,end=workdays[0],workdays[-1]
        label=f"{month_name} {start}" if start==end else f"{month_name} {start}-{end}"
        weeks.append({"weekNumber":wn,"workdays":workdays,"label":label})
    return {"year":year,"month":month_num,"monthName":month_name,"weeks":weeks,"weekCount":len(weeks)}

def get_month_workdays(year, month_num):
    days=monthrange(year,month_num)[1]; out=[]
    for day in range(1,days+1):
        d=date(year,month_num,day)
        if d.weekday()<5: out.append(d)
    return out

def get_day_name(d, short=True):
    idx=d.weekday(); names=["Mon","Tue","Wed","Thu","Fri","Sat","Sun"] if short else ["Monday","Tuesday","Wednesday","Thursday","Friday","Saturday","Sunday"]; return names[idx]

def group_consecutive(days):
    if not days: return []
    days=sorted(days); ranges=[]; start=end=days[0]
    for d in days[1:]:
        if d==end+1: end=d
        else: ranges.append({"start":start,"end":end}); start=end=d
    ranges.append({"start":start,"end":end}); return ranges


# ===========================================================================
# Fetch — download the Sheet tabs into a local .xlsx
#
# Always downloads fresh into a temp file (no on-disk cache). The team tab and
# month tab(s) are read **concurrently** (each `meta` call is ~2.6s of network);
# Sep/Sept candidates are probed in the same pool.
# ===========================================================================
DEFAULT_SHEET_ID = "1Vae2OUAdYT3pMAQLLSYcNRBFklJ2WybctNia6OjNK_g"
TEAM_TAB = "Team allocation 2026"
TEAM_RANGE = "A1:G200"
MONTH_RANGE = "A1:AC200"

def _month_tab_candidates(month, year):
    """Possible monthly tab titles, e.g. 'Jul 2026' (Sep also tries 'Sept 2026')."""
    short = get_month_name(get_month_number(month), True)
    aliases = ["Sep", "Sept"] if short == "Sep" else [short]
    return [f"{a} {year}" for a in aliases]

def _read_tab(sheet_id, tab, cell_range):
    """Read a tab via `meta google.sheets read`, returning a 2D list of cell values."""
    res = subprocess.run(
        ["meta", "google.sheets", "read", "--id", sheet_id,
         "--range", f"'{tab}'!{cell_range}", "-o", "json"],
        capture_output=True, text=True)
    out = (res.stdout or "").strip()
    start = out.find("[")
    if start == -1:
        raise RuntimeError(f"Could not read tab '{tab}'. {res.stderr.strip() or out}")
    data = json.loads(out[start:])
    if isinstance(data, dict):  # error payload came back as JSON
        raise RuntimeError(f"Could not read tab '{tab}': {data.get('message', data)}")
    return data

def _try_read_tab(sheet_id, tab, cell_range):
    """Like _read_tab but returns None on failure/empty (for concurrent candidate probing)."""
    try:
        rows = _read_tab(sheet_id, tab, cell_range)
        return rows or None
    except RuntimeError:
        return None

def _write_tab(wb, title, rows):
    ws = wb.create_sheet(title)
    for r, row in enumerate(rows, start=1):
        for c, val in enumerate(row, start=1):
            if val not in ("", None):
                ws.cell(r, c, val)
    return ws

def fetch_sheet(month, year, sheet_id=None, dest=None):
    """Download the team + monthly PTO tabs into a local .xlsx and return its path.

    Always downloads fresh (no on-disk cache); writes to a temp file that is
    overwritten each run.
    """
    sheet_id = sheet_id or DEFAULT_SHEET_ID
    dest = Path(dest) if dest else Path(tempfile.gettempdir()) / f"weekly-hours-sheet-{month}-{year}.xlsx"
    candidates = _month_tab_candidates(month, year)

    # team tab + month-tab candidates run concurrently (each meta call is ~2.6s of network)
    with ThreadPoolExecutor(max_workers=len(candidates) + 1) as ex:
        f_team = ex.submit(_read_tab, sheet_id, TEAM_TAB, TEAM_RANGE)
        f_month = {c: ex.submit(_try_read_tab, sheet_id, c, MONTH_RANGE) for c in candidates}
        team_rows = f_team.result()
        month_title = month_rows = None
        for cand in candidates:
            rows = f_month[cand].result()
            if rows:
                month_title, month_rows = cand, rows
                break

    if month_rows is None:
        raise RuntimeError(f"No monthly tab found for {month} {year} (tried {candidates}).")

    wb = Workbook()
    wb.remove(wb.active)  # drop the default empty sheet
    _write_tab(wb, TEAM_TAB, team_rows)
    _write_tab(wb, month_title, month_rows)
    dest.parent.mkdir(parents=True, exist_ok=True)
    wb.save(dest)
    return dest


# ===========================================================================
# Read PTO — monthly tab (Collab column, day columns, Backup/Notes)
# ===========================================================================
PTO_HEADER_ROW=8; DATA_START_ROW=9; COLLAB_COL=2; FIRST_DAY_COL=3; HOLIDAY_PURPLE="FFB4A7D6"

def is_purple(cell):
    fill=cell.fill
    if not fill or fill.patternType is None: return False
    fg=fill.fgColor
    if not fg or not fg.rgb: return False
    rgb=fg.rgb.upper(); return HOLIDAY_PURPLE in rgb or rgb in ("FF00FF","800080","FFC000FF","FF800080")

def get_sheet_names(month,year): aliases=["Sep","Sept"] if month=="Sep" else [month]; return [f"{m} {year}" for m in aliases]

def extract_cell_text(cell): v=cell.value; return "" if v is None else str(v).strip()

def read_pto(month,year,pto_path=None):
    if not pto_path:
        raise ValueError("pto_path (local xlsx fetched from the Google Sheet) is required")
    p=Path(pto_path)
    if not p.exists(): raise FileNotFoundError(f"PTO file not found at {p}")
    wb=load_workbook(p,data_only=True); ws=None
    for name in get_sheet_names(month,year):
        if name in wb.sheetnames: ws=wb[name]; break
    if not ws: raise ValueError(f"Sheet not found for {month} {year}")
    header=ws[PTO_HEADER_ROW]; backup_col=notes_col=None
    for idx,cell in enumerate(header,start=1):
        v=str(cell.value or "").strip().lower()
        if v=="backup": backup_col=idx
        if v=="notes": notes_col=idx
    if backup_col is None: backup_col=24
    if notes_col is None: notes_col=backup_col+1
    day_cols=[]
    for col_idx in range(FIRST_DAY_COL,backup_col):
        cell=ws.cell(PTO_HEADER_ROW,col_idx)
        try: day_num=int(cell.value)
        except: continue
        if 1<=day_num<=31: day_cols.append((col_idx,day_num,is_purple(cell)))
    result={}
    for row_idx in range(DATA_START_ROW,ws.max_row+1):
        employee_name=str(ws.cell(row_idx,COLLAB_COL).value or "").strip()
        if not employee_name: continue
        norm=normalize_value(employee_name); absences=[]
        for col_idx,day_num,is_hol in day_cols:
            if is_hol: absences.append({"dayNumber":day_num,"type":"H"}); continue
            v=ws.cell(row_idx,col_idx).value
            if v: absences.append({"dayNumber":day_num,"type":str(v).strip()})
        backups_text=extract_cell_text(ws.cell(row_idx,backup_col)); notes_text=extract_cell_text(ws.cell(row_idx,notes_col))
        result[norm]={"employeeName":employee_name,"absences":absences,"backupsText":backups_text,"notesText":notes_text}
    return result


# ===========================================================================
# Read Team — tab "Team allocation 2026"
# ===========================================================================
TEAM_HEADER_ROW = 1
# report field -> column header in the "Team allocation 2026" tab
COLS = {
    "team": "App",                         # product
    "module": "Feature",                   # pilar
    "qa_lead": "QA Lead",
    "job_role": "Role",                    # QA 1 / QA 2 / QA 3 -> Q1/Q2/Q3
    "short_name": "QA Analyst",            # match key vs PTO "Collab" + displayed name
    "full_name": "QA Analyst (Full Name)",
    "qa_3": "New QA3",
}

def _parse_team_rows(rows):
    """Parse the "Team allocation 2026" 2D cell rows into collaborator dicts."""
    if not rows: return []
    header = {}
    for i, h in enumerate(rows[0]):
        if h not in (None, ""): header[str(h).strip()] = i
    out = []
    for r in rows[1:]:
        def get(name):
            i = header.get(name)
            return str(r[i]).strip() if (i is not None and i < len(r) and r[i] not in (None, "")) else ""
        short = get(COLS["short_name"])
        if not short: continue
        rec = {field: get(h) for field, h in COLS.items()}
        # fields absent from the new Sheet, kept empty for downstream compatibility
        rec.update({"wave":"","meta_username":"","meta_tag":"","rate":"","softtek_username":""})
        out.append(rec)
    return out

def read_team(team_path=None):
    if not team_path:
        raise ValueError("team_path (local xlsx fetched from the Google Sheet) is required")
    p = Path(team_path)
    if not p.exists(): raise FileNotFoundError(f"Team file not found at {p}")
    wb = load_workbook(p, data_only=True); ws = wb[TEAM_TAB]
    rows = [[c.value for c in row] for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column)]
    return _parse_team_rows(rows)

def fetch_team_roster(sheet_id=None):
    """Read just the "Team allocation 2026" tab (one network call) → collaborator list.
    Used to validate rates before a month is known."""
    rows = _read_tab(sheet_id or DEFAULT_SHEET_ID, TEAM_TAB, TEAM_RANGE)
    return _parse_team_rows(rows)


# ===========================================================================
# Build report — combine PTO + team into Q1/Q2/Q3 structures
# ===========================================================================
ABSENCE_TYPES = {"pto","pto(pa)","ml","ml(pa)"}
# "O" (gray) = collaborator is OUTSIDE the project those days. It's not a PTO/ML
# absence, but it IS coverable: a backup credited on the covered person's "O" days
# (billed at that person's rate), just like a real PTO/ML day.
OUT_TYPES = {"o"}
COVERABLE_TYPES = ABSENCE_TYPES | OUT_TYPES

def parse_backups(text):
    if not text.strip():
        return []
    out=[]
    for line in text.split("\n"):
        line = re.sub(r"\(.*?\)","",line).strip()
        if not line:
            continue
        m = re.match(r"^[A-Za-zÀ-ÿ\s]+", line)
        if not m:
            continue
        name = m.group(0).strip()
        norm = normalize_value(name)
        rest = line[len(m.group(0)):].strip()
        for rng in rest.split(","):
            rng=rng.strip()
            if not rng:
                continue
            parts = [p.strip() for p in rng.split("-")]
            try:
                start=int(parts[0]); end=int(parts[1]) if len(parts)>1 else start
            except:
                continue
            out.append({"backupName":norm,"originalBackupName":name,"startDay":start,"endDay":end})
    return out

def parse_team_changes(notes):
    changes=[]
    for m in re.finditer(r"TEAM:\s*(.+?)\s*\|\s*(.+?)\s+(\d+)-(\d+)", notes):
        changes.append({"product":m.group(1).strip(),"pilar":m.group(2).strip(),"fromDay":int(m.group(3)),"toDay":int(m.group(4))})
    return changes

def get_bench_days(team_changes, year, month):
    bench=set()
    for ch in team_changes:
        if "bench" not in ch["product"].lower():
            continue
        for day in range(ch["fromDay"], ch["toDay"]+1):
            try:
                d=date(year,month,day)
            except:
                continue
            if d.month!=month or d.weekday()>=5:
                continue
            bench.add(day)
    return bench

def build_comments(weeks):
    mp=defaultdict(list)
    for w in weeks:
        for a in w["absences"]:
            t=a["type"]
            if normalize_value(t) in ("h","holiday"):
                continue
            mp[t].append(a["dayNumber"])
    comments=[]
    for typ, days in mp.items():
        ranges=group_consecutive(sorted(days))
        rng_txt=",".join(f"{r['start']}" if r["start"]==r["end"] else f"{r['start']}-{r['end']}" for r in ranges)
        comments.append(f"• {typ}: {rng_txt}")
    return "\n".join(comments)

def build_team_note(changes):
    if not changes:
        return ""
    return "\n".join(f"• Team change: {c['product']} | {c['pilar']} (days {c['fromDay']}-{c['toDay']})" for c in changes)

def build_report(pto_data, team_data, month, year):
    master={normalize_value(c["short_name"]):c for c in team_data if c["short_name"]}
    month_num=get_month_number(month)
    ctx=build_month_context(year, month)
    workdays=get_month_workdays(year, month_num)
    q1,q2,q3=[],[],[]
    for norm_name,data in pto_data.items():
        emp=master.get(norm_name)
        if not emp:
            print(f"WARN {data['employeeName']} not in team allocation, skipping")
            continue
        if "bench" in emp["team"].lower():
            continue
        weekly={}
        for d in workdays:
            wn=get_week_of_month(d)
            weekly.setdefault(wn,{"weekNumber":wn,"absences":[],"totalAbsenceHours":0,"totalWorkedHours":0,"backups":[]})
            weekly[wn]["totalWorkedHours"]+=8
        absence_map={}; holiday_days=set()
        for a in data["absences"]:
            nt=normalize_value(a["type"])
            absence_map[a["dayNumber"]]=a["type"]
            if nt in ("h","holiday"):
                holiday_days.add(a["dayNumber"])
        for a in data["absences"]:
            d=date(year,month_num,a["dayNumber"]); wn=get_week_of_month(d)
            if wn not in weekly: continue
            w=weekly[wn]; nt=normalize_value(a["type"])
            if nt in ("h","holiday"):
                w["absences"].append(a); w["totalWorkedHours"]-=8; continue
            if nt in ABSENCE_TYPES:
                w["absences"].append(a); w["totalAbsenceHours"]+=8; w["totalWorkedHours"]-=8; continue
            w["totalWorkedHours"]-=8
        team_changes=parse_team_changes(data["notesText"])
        bench_days=get_bench_days(team_changes, year, month_num)
        bench_hrs=0
        for day in bench_days:
            if day in absence_map: continue
            d=date(year,month_num,day); wn=get_week_of_month(d)
            w=weekly.get(wn)
            if w: w["totalWorkedHours"]-=8; bench_hrs+=8
        backup_ranges=parse_backups(data["backupsText"])
        absence_days_set={a["dayNumber"] for a in data["absences"] if normalize_value(a["type"]) in COVERABLE_TYPES}
        # A backup covers the person's coverable days only (PTO/ML or "O" out-of-project).
        # Ranges may span weekends/holidays (e.g. "2-6" over a holiday on the 3rd) — those
        # days aren't worked, so warn on any unexpected weekday but never credit or crash on them.
        for br in backup_ranges:
            for day in range(br["startDay"], br["endDay"]+1):
                try: d=date(year,month_num,day)
                except: continue
                if d.weekday()>=5 or day in holiday_days: continue
                if day not in absence_days_set:
                    print(f"WARN {emp['short_name']}: backup {br['originalBackupName']} on day {day} without PTO/ML/O, skipping that day")
        bc={}
        for br in backup_ranges:
            be=master.get(br["backupName"])
            if not be:
                print(f"WARN backup {br['originalBackupName']} not found, skipping"); continue
            for day in range(br["startDay"], br["endDay"]+1):
                if day not in absence_days_set: continue  # only cover PTO/ML/O days (excludes weekends & holidays)
                try: d=date(year,month_num,day)
                except: continue
                if d.weekday()>=5: continue
                wn=get_week_of_month(d)
                bc.setdefault(br["backupName"],{}).setdefault(wn,{"backupName":be["short_name"],"totalHours":0,"workdays":set(),"employeeData":be})
                bc[br["backupName"]][wn]["totalHours"]+=8
                bc[br["backupName"]][wn]["workdays"].add(day)
        for bn,wm in bc.items():
            for wn,cd in wm.items():
                if wn not in weekly: continue
                w=weekly[wn]; wds=sorted(cd["workdays"])
                w["backups"].append({"name":cd["employeeData"]["short_name"],"role":cd["employeeData"]["job_role"],"wave":cd["employeeData"]["wave"],"product":cd["employeeData"]["team"],"pilar":cd["employeeData"]["module"],"startDay":min(wds),"endDay":max(wds),"hours":cd["totalHours"],"workdays":wds})
        all_weeks=sorted(weekly.values(),key=lambda x:x["weekNumber"])
        week_details=[]
        for w in all_weeks:
            wds=[d for d in workdays if get_week_of_month(d)==w["weekNumber"]]
            days=[]
            for d in wds:
                dn=d.day; raw=absence_map.get(dn)
                if not raw:
                    if dn in bench_days:
                        days.append({"dayNumber":dn,"dayName":get_day_name(d),"status":"Bench","hours":0})
                    else:
                        days.append({"dayNumber":dn,"dayName":get_day_name(d),"status":"","hours":8})
                    continue
                nt=normalize_value(raw)
                if nt in ("h","holiday"):
                    days.append({"dayNumber":dn,"dayName":get_day_name(d),"status":"H","hours":0})
                elif nt in ABSENCE_TYPES:
                    days.append({"dayNumber":dn,"dayName":get_day_name(d),"status":raw,"hours":0})
                elif nt in OUT_TYPES:
                    days.append({"dayNumber":dn,"dayName":get_day_name(d),"status":"O","hours":0})
                else:
                    days.append({"dayNumber":dn,"dayName":get_day_name(d),"status":"","hours":0})
            meta=next((x for x in ctx["weeks"] if x["weekNumber"]==w["weekNumber"]),None)
            week_details.append({"weekNumber":w["weekNumber"],"label":meta["label"] if meta else f"Week {w['weekNumber']}","totalHours":w["totalWorkedHours"],"days":days})
        week_hours=[w["totalWorkedHours"] for w in all_weeks]
        total_abs=sum(w["totalAbsenceHours"] for w in all_weeks)
        total_work=sum(w["totalWorkedHours"] for w in all_weeks)
        comments="\n".join(filter(None,[build_team_note(team_changes), build_comments(all_weeks)]))
        employee={"fullName":emp["short_name"],"role":emp["job_role"],"wave":emp["wave"],"product":emp["team"],"pilar":emp["module"],"tag":f"{emp['short_name']} - {emp['team']} - {emp['module']}","weekHours":week_hours,"weekDetails":week_details,"absHrs":total_abs,"workHrs":total_work,"benchHrs":bench_hrs,"comments":comments,"isBench":False,"isBackup":False,"isSeparator":False,"hasIncompleteData":False}
        role=emp["job_role"]
        target = q1 if "QA 1" in role else q2 if "QA 2" in role else q3 if "QA 3" in role else None
        if target is not None:
            target.append(employee)
            backup_map={}
            for wi,w in enumerate(all_weeks):
                for b in w["backups"]:
                    key=b["name"]
                    if key not in backup_map:
                        backup_map[key]={"name":b["name"],"role":b["role"],"wave":b["wave"],"product":b["product"],"pilar":b["pilar"],"weeklyHours":[0]*ctx["weekCount"],"totalHours":0,"allWorkdays":[],"coveredDaysByWeek":{}}
                    bm=backup_map[key]
                    bm["weeklyHours"][wi]+=b["hours"]; bm["totalHours"]+=b["hours"]
                    bm["coveredDaysByWeek"].setdefault(wi,set()).update(b["workdays"])
                    for day in b["workdays"]:
                        if day not in bm["allWorkdays"]: bm["allWorkdays"].append(day)
            for bm in backup_map.values():
                wds=[]
                for wi in range(ctx["weekCount"]):
                    wds_w=[d for d in workdays if get_week_of_month(d)==all_weeks[wi]["weekNumber"]] if wi < len(all_weeks) else []
                    covered=bm["coveredDaysByWeek"].get(wi,set())
                    wds.append({"weekNumber": all_weeks[wi]["weekNumber"] if wi < len(all_weeks) else wi+1,
                                "label": week_details[wi]["label"] if wi < len(week_details) else f"Week {wi+1}",
                                "totalHours": bm["weeklyHours"][wi],
                                "days":[{"dayNumber":d.day,"dayName":get_day_name(d),"status":"","hours":8 if d.day in covered else 0} for d in wds_w]})
                ranges=group_consecutive(sorted(bm["allWorkdays"]))
                rng_txt=",".join(f"{r['start']}-{r['end']}" if r["end"]-r["start"]+1>=4 else ",".join(str(x) for x in range(r["start"],r["end"]+1)) for r in ranges)
                comment=f"• Covered {emp['short_name']} {'day' if len(bm['allWorkdays'])==1 else 'days'} {rng_txt}"
                backup_row={"fullName":f"↳ {bm['name']}","role":bm["role"],"wave":bm["wave"],"product":bm["product"],"pilar":bm["pilar"],"tag":f"{bm['name']} - {bm['product']} - {bm['pilar']}","weekHours":bm["weeklyHours"],"weekDetails":wds,"absHrs":-1,"workHrs":bm["totalHours"],"benchHrs":-1,"comments":comment,"isBench":False,"isBackup":True,"isSeparator":False,"hasIncompleteData":False,"coveringFor":emp["short_name"]}
                target.append(backup_row)
    def sort_emps(lst):
        principals=[e for e in lst if not e["isBackup"]]
        backups=[e for e in lst if e["isBackup"]]
        principals.sort(key=lambda x:(x["product"],x["fullName"]))
        res=[]
        for p in principals:
            res.append(p); res.extend([b for b in backups if b.get("coveringFor")==p["fullName"]])
        return res
    return {"q1":sort_emps(q1),"q2":sort_emps(q2),"q3":sort_emps(q3)}


# ===========================================================================
# Export — formatted, color-coded Excel workbook
# ===========================================================================
# Header/section fills use the original color-by-group scheme (identity=navy,
# weeks=gold, Abs+Work=green, Comments+Tag=purple). The dark DIVIDER lines that
# frame the column groups are kept on top of the colors.
NAVY="2C3E7B"; GOLD="7B5E1A"; GREEN="1B7A3D"; PURPLE="5B2D8B"; AMBER="F59E0B"
BACKUP_CREAM="FEF3C7"; TOTALS_GREEN="B6D7A8"; EMERGENCY_RED="FECACA"
WHITE="FFFFFF"; HEADER_TEXT="FFFFFF"
GRID_GRAY="D9D9D9"      # light internal gridlines
DIVIDER_BLACK="000000"  # dark lines that separate the column groups

def fill(argb): return PatternFill(start_color=argb, end_color=argb, fill_type="solid")
def thin(): return Side(style="thin", color=GRID_GRAY)
def divider(): return Side(style="medium", color=DIVIDER_BLACK)

def section_color(col, week_count, weeks_detail_cols):
    if col<=5: return NAVY
    gold_end=5+weeks_detail_cols   # weeks area = Mon-Fri days + weekly Hrs (no per-week label col)
    if col<=gold_end: return GOLD
    if col<=gold_end+2: return GREEN
    return PURPLE

def write_weekly_sheet(wb, data, title="Weekly Hours"):
    """Add a formatted 'Weekly Hours' worksheet to an existing workbook."""
    employees=[]
    for label, lst in [("Q1",data["q1"]),("Q2",data["q2"]),("Q3",data["q3"])]:
        if lst:
            employees.append({"fullName":label,"isSeparator":True})
            employees.extend(lst)
    first = next((e for e in employees if not e.get("isSeparator")), None)
    if not first:
        print("No weekly data"); return None
    week_count=len(first["weekHours"])
    weeks_detail = first.get("weekDetails",[])
    # Each week block = Mon-Fri day columns + one weekly Hrs (sum) column.
    weeks_detail_cols = sum(len(w["days"])+1 for w in weeks_detail)
    total_cols = 5 + weeks_detail_cols + 4
    abs_col = 6 + weeks_detail_cols
    comments_col = abs_col + 2

    ws=wb.create_sheet(title)
    widths=[25,10,10,15,15]
    for w in weeks_detail:
        for _ in w["days"]: widths.append(8)
        widths.append(8)   # weekly Hrs (sum) column
    widths.extend([10,10,40,35])
    for i,w in enumerate(widths,1): ws.column_dimensions[get_column_letter(i)].width=w

    header_font=Font(bold=True,color=HEADER_TEXT,size=10)
    center=Alignment(horizontal="center",vertical="center")
    wrap=Alignment(horizontal="center",vertical="center",wrap_text=True)

    # Dark divider lines sit on the RIGHT edge of these columns, framing the groups:
    # identity (1-5) | each Week block | Abs+Work | Comments | Tag.
    divider_cols={5}
    c=6
    for w in weeks_detail:
        c+=len(w["days"])+1
        divider_cols.add(c-1)        # right edge of this week block (last one == weeks|Abs)
    divider_cols.add(abs_col+1)      # after Work Hrs (Abs+Work | Comments)
    divider_cols.add(comments_col)   # after Comments (Comments | Tag)
    divider_cols.add(total_cols)     # outer right edge

    def make_border(col, top_dark=False, bottom_dark=False):
        return Border(
            left=divider() if (col==1 or (col-1) in divider_cols) else thin(),
            right=divider() if col in divider_cols else thin(),
            top=divider() if top_dark else thin(),
            bottom=divider() if bottom_dark else thin())

    for c in range(1,total_cols+1):
        cell=ws.cell(1,c); sc=section_color(c,week_count,weeks_detail_cols)
        cell.fill=fill(sc); cell.font=header_font; cell.alignment=center; cell.border=make_border(c, top_dark=True)
    # "Week N" spans the full width of its week block (days + weekly Hrs).
    col_idx=6
    for i,w in enumerate(weeks_detail):
        block_len=len(w["days"])+1
        ws.cell(1,col_idx).value=f"Week {i+1}"
        ws.merge_cells(start_row=1,start_column=col_idx,end_row=1,end_column=col_idx+block_len-1)
        col_idx+=block_len

    headers=["Employee","Role","Wave","Product","Pilar"]
    for w in weeks_detail:
        for d in w["days"]:
            headers.append(f"{d['dayName']} {d['dayNumber']}")
        headers.append("Hrs")
    headers.extend(["Abs Hrs","Work Hrs","Comments","Tag"])
    for i,h in enumerate(headers,1):
        cell=ws.cell(2,i,h); sc=section_color(i,week_count,weeks_detail_cols)
        cell.fill=fill(sc); cell.font=header_font; cell.alignment=center; cell.border=make_border(i, bottom_dark=True)

    work_refs={}  # id(emp) -> "'Weekly Hours'!<col><row>" of its Work Hrs cell (for live Invoice references)
    tag_refs={}   # id(emp) -> "'Weekly Hours'!<col><row>" of its Tag cell (for the Invoice DESCRIPTION formula)
    work_col_letter=get_column_letter(abs_col+1)
    tag_col_letter=get_column_letter(comments_col+1)
    row_idx=3
    for emp in employees:
        if emp.get("isSeparator"):
            ws.cell(row_idx,1,emp["fullName"])
            for c in range(1,total_cols+1):
                cell=ws.cell(row_idx,c); cell.fill=fill(AMBER); cell.font=Font(bold=True,color=WHITE); cell.alignment=wrap; cell.border=make_border(c)
            row_idx+=1; continue
        is_backup=emp.get("isBackup",False)
        is_emergency=not is_backup and emp.get("absHrs",0)!=-1 and emp["absHrs"]>80
        ws.cell(row_idx,1,emp["fullName"]); ws.cell(row_idx,2,emp["role"]); ws.cell(row_idx,3,emp["wave"]); ws.cell(row_idx,4,emp["product"]); ws.cell(row_idx,5,emp["pilar"])
        col=6
        hrs_cells=[]   # the per-week Hrs cells for this row — Work Hrs sums them
        for wi,hrs in enumerate(emp["weekHours"]):
            wd=emp["weekDetails"][wi] if wi < len(emp["weekDetails"]) else {"days":[]}
            for d in wd["days"]:
                cell=ws.cell(row_idx,col); cell.alignment=center
                if is_backup:
                    cell.value = 8 if d["hours"]>0 else ""
                else:
                    cell.value = d["status"] or ""
                st=d["status"]
                if st=="PTO": cell.fill=fill("3B82F6"); cell.font=Font(color="FFFFFF")
                elif st in ("PTO(PA)","ML","ML(PA)"): cell.fill=fill("FACC15"); cell.font=Font(color="000000")
                elif st=="H": cell.fill=fill("F3E8FF"); cell.font=Font(color="A855F7")
                elif st=="Bench": cell.fill=fill("E5E7EB"); cell.font=Font(color="6B7280")
                elif st=="O": cell.fill=fill("D1D5DB"); cell.font=Font(color="4B5563")  # out-of-project (gray)
                col+=1
            ws.cell(row_idx,col,hrs).alignment=center   # weekly Hrs (editable) column
            hrs_cells.append(f"{get_column_letter(col)}{row_idx}"); col+=1
        ws.cell(row_idx,abs_col, "N/A" if emp["absHrs"]==-1 else emp["absHrs"]).alignment=center
        # Work Hrs = SUM of the weekly Hrs cells, so editing a collaborator's weekly
        # hours flows into Work Hrs → and into the linked Invoice QTY/AMOUNT.
        work_value=f"=SUM({','.join(hrs_cells)})" if hrs_cells else emp["workHrs"]
        ws.cell(row_idx,abs_col+1, work_value).alignment=center
        work_refs[id(emp)]=f"'{title}'!{work_col_letter}{row_idx}"
        ws.cell(row_idx,comments_col, emp["comments"]).alignment=wrap
        ws.cell(row_idx,comments_col+1, emp["tag"]).alignment=center
        tag_refs[id(emp)]=f"'{title}'!{tag_col_letter}{row_idx}"
        for c in range(1,total_cols+1):
            cell=ws.cell(row_idx,c)
            cell.border=make_border(c)
            if not cell.alignment.horizontal: cell.alignment=center
            if is_backup: cell.fill=fill(BACKUP_CREAM)
            elif is_emergency and (not cell.fill.start_color.index or cell.fill.start_color.index=='00000000'):
                cell.fill=fill(EMERGENCY_RED)
        row_idx+=1

    ws.cell(row_idx,1,"Totals").font=Font(bold=True)
    data_start=3; data_end=row_idx-1
    col=6
    for wi in range(week_count):
        ndays=len(weeks_detail[wi]["days"]) if wi < len(weeks_detail) else 0
        hrs_col=col+ndays   # the weekly Hrs (sum) column at the end of this week block
        letter=get_column_letter(hrs_col)
        ws.cell(row_idx,hrs_col).value=f"=SUM({letter}{data_start}:{letter}{data_end})"
        ws.cell(row_idx,hrs_col).font=Font(bold=True); ws.cell(row_idx,hrs_col).alignment=center
        col+=ndays+1
    abs_letter=get_column_letter(abs_col); ws.cell(row_idx,abs_col).value=f"=SUM({abs_letter}{data_start}:{abs_letter}{data_end})"; ws.cell(row_idx,abs_col).font=Font(bold=True); ws.cell(row_idx,abs_col).alignment=center
    work_letter=get_column_letter(abs_col+1); ws.cell(row_idx,abs_col+1).value=f"=SUM({work_letter}{data_start}:{work_letter}{data_end})"; ws.cell(row_idx,abs_col+1).font=Font(bold=True); ws.cell(row_idx,abs_col+1).alignment=center
    total_backup=sum(e["workHrs"] for e in employees if e.get("isBackup") and e.get("workHrs",0)>0)
    ws.cell(row_idx,comments_col,f"Total Backup Hours: {total_backup}").font=Font(bold=True); ws.cell(row_idx,comments_col).alignment=center
    for c in range(1,total_cols+1):
        cell=ws.cell(row_idx,c); cell.fill=fill(TOTALS_GREEN); cell.border=make_border(c, top_dark=True, bottom_dark=True)

    ws.freeze_panes="B3"
    return work_refs, tag_refs

def export_excel(data, output_path: Path):
    """Standalone: write just the Weekly Hours sheet to its own workbook."""
    wb=Workbook(); wb.remove(wb.active)
    if write_weekly_sheet(wb, data) is None:
        return
    output_path.parent.mkdir(parents=True,exist_ok=True)
    wb.save(output_path)
    print(f"Saved {output_path}")
