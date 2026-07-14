"""Export weekly hours report to Excel with formatting."""
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Font, Alignment
from openpyxl.utils import get_column_letter

NAVY="2C3E7B"; GOLD="7B5E1A"; GREEN="1B7A3D"; PURPLE="5B2D8B"; AMBER="F59E0B"
BACKUP_CREAM="FEF3C7"; TOTALS_GREEN="B6D7A8"; EMERGENCY_RED="FECACA"; WHITE="FFFFFF"; BORDER_GRAY="444444"

def fill(argb): return PatternFill(start_color=argb, end_color=argb, fill_type="solid")
def thin(): return Side(style="thin", color=BORDER_GRAY)

def section_color(col, week_count, weeks_detail_cols):
    if col<=5: return NAVY
    gold_end=5+week_count+weeks_detail_cols
    if col<=gold_end: return GOLD
    if col<=gold_end+2: return GREEN
    return PURPLE

def export_excel(data, output_path: Path):
    employees=[]
    for label, lst in [("Q1",data["q1"]),("Q2",data["q2"]),("Q3",data["q3"])]:
        if lst:
            employees.append({"fullName":label,"isSeparator":True})
            employees.extend(lst)
    first = next((e for e in employees if not e.get("isSeparator")), None)
    if not first:
        print("No data"); return
    week_count=len(first["weekHours"])
    weeks_detail = first.get("weekDetails",[])
    weeks_detail_cols = sum(len(w["days"])+1 for w in weeks_detail)
    total_cols = 5 + week_count + weeks_detail_cols + 4
    abs_col = 6 + week_count + weeks_detail_cols
    comments_col = abs_col + 2

    wb=Workbook(); ws=wb.active; ws.title="Weekly Hours"
    widths=[25,10,10,15,15]
    for w in weeks_detail:
        widths.append(12)
        for _ in w["days"]: widths.append(8)
        widths.append(8)
    widths.extend([10,10,40,35])
    for i,w in enumerate(widths,1): ws.column_dimensions[get_column_letter(i)].width=w

    header_font=Font(bold=True,color=WHITE,size=10)
    center=Alignment(horizontal="center",vertical="center")
    wrap=Alignment(horizontal="center",vertical="center",wrap_text=True)
    border=Border(left=thin(),right=thin(),top=thin(),bottom=thin())

    for c in range(1,total_cols+1):
        cell=ws.cell(1,c); sc=section_color(c,week_count,weeks_detail_cols)
        cell.fill=fill(sc); cell.font=header_font; cell.alignment=center; cell.border=border
    col_idx=6
    for i,w in enumerate(weeks_detail):
        ws.cell(1,col_idx).value=f"Week {i+1}"
        col_idx+=1+len(w["days"])+1

    headers=["Employee","Role","Wave","Product","Pilar"]
    for w in weeks_detail:
        headers.append(w["label"])
        for d in w["days"]:
            headers.append(f"{d['dayName']} {d['dayNumber']}")
        headers.append("Hrs")
    headers.extend(["Abs Hrs","Work Hrs","Comments","Tag"])
    for i,h in enumerate(headers,1):
        cell=ws.cell(2,i,h); sc=section_color(i,week_count,weeks_detail_cols)
        cell.fill=fill(sc); cell.font=header_font; cell.alignment=center; cell.border=border

    row_idx=3
    for emp in employees:
        if emp.get("isSeparator"):
            ws.cell(row_idx,1,emp["fullName"])
            for c in range(1,total_cols+1):
                cell=ws.cell(row_idx,c); cell.fill=fill(AMBER); cell.font=Font(bold=True,color=WHITE); cell.alignment=wrap; cell.border=border
            row_idx+=1; continue
        is_backup=emp.get("isBackup",False)
        is_emergency=not is_backup and emp.get("absHrs",0)!=-1 and emp["absHrs"]>80
        ws.cell(row_idx,1,emp["fullName"]); ws.cell(row_idx,2,emp["role"]); ws.cell(row_idx,3,emp["wave"]); ws.cell(row_idx,4,emp["product"]); ws.cell(row_idx,5,emp["pilar"])
        col=6
        for wi,hrs in enumerate(emp["weekHours"]):
            ws.cell(row_idx,col,hrs).alignment=center; col+=1
            wd=emp["weekDetails"][wi] if wi < len(emp["weekDetails"]) else {"days":[]}
            for d in wd["days"]:
                cell=ws.cell(row_idx,col); cell.alignment=center
                if is_backup:
                    cell.value = 8 if d["hours"]>0 else ""
                else:
                    cell.value = d["status"] or ""
                st=d["status"]
                if st=="PTO": cell.fill=fill("3B82F6"); cell.font=Font(color="FFFFFF")
                elif st in ("PTO(PA)","ML","ML(PA)"): cell.fill=fill("FACC15"); cell.font=Font(color="000000")
                elif st=="H": cell.fill=fill("F3E8FF"); cell.font=Font(color="A855F7")
                elif st=="Bench": cell.fill=fill("E5E7EB"); cell.font=Font(color="6B7280")
                col+=1
            ws.cell(row_idx,col,hrs).alignment=center; col+=1
        ws.cell(row_idx,abs_col, "N/A" if emp["absHrs"]==-1 else emp["absHrs"]).alignment=center
        ws.cell(row_idx,abs_col+1, emp["workHrs"]).alignment=center
        ws.cell(row_idx,comments_col, emp["comments"]).alignment=wrap
        ws.cell(row_idx,comments_col+1, emp["tag"]).alignment=center
        for c in range(1,total_cols+1):
            cell=ws.cell(row_idx,c)
            if not cell.border.left.style: cell.border=border
            if not cell.alignment.horizontal: cell.alignment=center
            if is_backup: cell.fill=fill(BACKUP_CREAM)
            elif is_emergency and (not cell.fill.start_color.index or cell.fill.start_color.index=='00000000'):
                cell.fill=fill(EMERGENCY_RED)
        row_idx+=1

    ws.cell(row_idx,1,"Totals").font=Font(bold=True)
    data_start=3; data_end=row_idx-1
    col=6
    for wi in range(week_count):
        letter=get_column_letter(col)
        ws.cell(row_idx,col).value=f"=SUM({letter}{data_start}:{letter}{data_end})"
        ws.cell(row_idx,col).font=Font(bold=True); ws.cell(row_idx,col).alignment=center
        col+=1+len(weeks_detail[wi]["days"])+1 if wi < len(weeks_detail) else 1
    abs_letter=get_column_letter(abs_col); ws.cell(row_idx,abs_col).value=f"=SUM({abs_letter}{data_start}:{abs_letter}{data_end})"; ws.cell(row_idx,abs_col).font=Font(bold=True); ws.cell(row_idx,abs_col).alignment=center
    work_letter=get_column_letter(abs_col+1); ws.cell(row_idx,abs_col+1).value=f"=SUM({work_letter}{data_start}:{work_letter}{data_end})"; ws.cell(row_idx,abs_col+1).font=Font(bold=True); ws.cell(row_idx,abs_col+1).alignment=center
    total_backup=sum(e["workHrs"] for e in employees if e.get("isBackup") and e.get("workHrs",0)>0)
    ws.cell(row_idx,comments_col,f"Total Backup Hours: {total_backup}").font=Font(bold=True); ws.cell(row_idx,comments_col).alignment=center
    for c in range(1,total_cols+1):
        cell=ws.cell(row_idx,c); cell.fill=fill(TOTALS_GREEN); cell.border=border

    ws.freeze_panes="B3"
    output_path.parent.mkdir(parents=True,exist_ok=True)
    wb.save(output_path)
    print(f"Saved {output_path}")
