from pathlib import Path
from openpyxl import load_workbook
from utils import normalize_value
HEADER_ROW=8; DATA_START_ROW=9; COLLAB_COL=2; FIRST_DAY_COL=3; HOLIDAY_PURPLE="FFB4A7D6"
def is_purple(cell):
    fill=cell.fill
    if not fill or fill.patternType is None: return False
    fg=fill.fgColor
    if not fg or not fg.rgb: return False
    rgb=fg.rgb.upper(); return HOLIDAY_PURPLE in rgb or rgb in ("FF00FF","800080","FFC000FF","FF800080")
def get_sheet_names(month,year): aliases=["Sep","Sept"] if month=="Sep" else [month]; return [f"{m} {year}" for m in aliases]
def extract_cell_text(cell): v=cell.value; return "" if v is None else str(v).strip()
def read_pto(month,year,pto_path=None):
    if not pto_path:
        raise ValueError("pto_path (local xlsx fetched from the Google Sheet) is required")
    p=Path(pto_path)
    if not p.exists(): raise FileNotFoundError(f"PTO file not found at {p}")
    wb=load_workbook(p,data_only=True); ws=None
    for name in get_sheet_names(month,year):
        if name in wb.sheetnames: ws=wb[name]; break
    if not ws: raise ValueError(f"Sheet not found for {month} {year}")
    header=ws[HEADER_ROW]; backup_col=notes_col=None
    for idx,cell in enumerate(header,start=1):
        v=str(cell.value or "").strip().lower()
        if v=="backup": backup_col=idx
        if v=="notes": notes_col=idx
    if backup_col is None: backup_col=24
    if notes_col is None: notes_col=backup_col+1
    day_cols=[]
    for col_idx in range(FIRST_DAY_COL,backup_col):
        cell=ws.cell(HEADER_ROW,col_idx)
        try: day_num=int(cell.value)
        except: continue
        if 1<=day_num<=31: day_cols.append((col_idx,day_num,is_purple(cell)))
    result={}
    for row_idx in range(DATA_START_ROW,ws.max_row+1):
        employee_name=str(ws.cell(row_idx,COLLAB_COL).value or "").strip()
        if not employee_name: continue
        norm=normalize_value(employee_name); absences=[]
        for col_idx,day_num,is_hol in day_cols:
            if is_hol: absences.append({"dayNumber":day_num,"type":"H"}); continue
            v=ws.cell(row_idx,col_idx).value
            if v: absences.append({"dayNumber":day_num,"type":str(v).strip()})
        backups_text=extract_cell_text(ws.cell(row_idx,backup_col)); notes_text=extract_cell_text(ws.cell(row_idx,notes_col))
        result[norm]={"employeeName":employee_name,"absences":absences,"backupsText":backups_text,"notesText":notes_text}
    return result
