"""Fetch collaborator + PTO data from the Google Sheet into a local .xlsx.

Everything (team allocation AND monthly PTO) lives in a single Google Sheet
("PTO Tracker Softtek"). We pull the needed tabs with `meta google.sheets read`
(works in any authenticated environment) and assemble a local workbook that
read_team.py and read_pto.py consume with openpyxl.

Performance:
- The two tab reads run **in parallel** (each `meta` call is ~2.6s of network).
- The month tab candidates (Sep/Sept) are probed **concurrently**, so resolving
  the real name never adds a sequential round-trip.
- A local **cache** avoids re-downloading: within CACHE_FRESH_SECONDS the cached
  file is reused with no network; after that we compare the Sheet's Drive
  `modifiedTime` and only re-download if it actually changed.
"""
import json
import os
import subprocess
import time
from pathlib import Path
from concurrent.futures import ThreadPoolExecutor
from openpyxl import Workbook
from utils import get_month_name, get_month_number

DEFAULT_SHEET_ID = "1Vae2OUAdYT3pMAQLLSYcNRBFklJ2WybctNia6OjNK_g"
TEAM_TAB = "Team allocation 2026"
TEAM_RANGE = "A1:G200"
MONTH_RANGE = "A1:AC200"
CACHE_FRESH_SECONDS = 120  # within this window: reuse cache with zero network


def _month_tab_candidates(month, year):
    """Possible monthly tab titles, e.g. 'Jul 2026' (Sep also tries 'Sept 2026')."""
    short = get_month_name(get_month_number(month), True)
    aliases = ["Sep", "Sept"] if short == "Sep" else [short]
    return [f"{a} {year}" for a in aliases]


def _read_tab(sheet_id, tab, cell_range):
    """Read a tab via `meta google.sheets read`, returning a 2D list of cell values."""
    res = subprocess.run(
        ["meta", "google.sheets", "read", "--id", sheet_id,
         "--range", f"'{tab}'!{cell_range}", "-o", "json"],
        capture_output=True, text=True)
    out = (res.stdout or "").strip()
    start = out.find("[")
    if start == -1:
        raise RuntimeError(f"Could not read tab '{tab}'. {res.stderr.strip() or out}")
    data = json.loads(out[start:])
    if isinstance(data, dict):  # error payload came back as JSON
        raise RuntimeError(f"Could not read tab '{tab}': {data.get('message', data)}")
    return data


def _try_read_tab(sheet_id, tab, cell_range):
    """Like _read_tab but returns None on failure/empty (for concurrent candidate probing)."""
    try:
        rows = _read_tab(sheet_id, tab, cell_range)
        return rows or None
    except RuntimeError:
        return None


def _sheet_modified_time(sheet_id):
    """Return the Sheet's modifiedTime (ISO string) or None.

    Uses `google.sheets describe` (~3s) rather than `google.drive.file describe`,
    which is ~20s for this file.
    """
    res = subprocess.run(
        ["meta", "google.sheets", "describe", "--id", sheet_id, "-o", "json"],
        capture_output=True, text=True)
    out = (res.stdout or "").strip()
    start = out.find("{")
    if start == -1:
        return None
    try:
        return json.loads(out[start:]).get("modifiedTime")
    except Exception:
        return None


def _write_tab(wb, title, rows):
    ws = wb.create_sheet(title)
    for r, row in enumerate(rows, start=1):
        for c, val in enumerate(row, start=1):
            if val not in ("", None):
                ws.cell(r, c, val)
    return ws


def _load_meta(meta_path):
    try:
        return json.loads(meta_path.read_text())
    except Exception:
        return {}


def fetch_sheet(month, year, sheet_id=None, dest=None, force=False):
    """Download the team + monthly PTO tabs into a local .xlsx and return its path.

    Uses the local cache when possible (see module docstring). Pass force=True to
    always re-download.
    """
    sheet_id = sheet_id or DEFAULT_SHEET_ID
    dest = Path(dest) if dest else Path(__file__).parent / "cache" / f"sheet-{month}-{year}.xlsx"
    meta_path = dest.with_suffix(".meta.json")
    candidates = _month_tab_candidates(month, year)

    # --- cache reuse ---
    current_mod = None
    if not force and dest.exists():
        age = time.time() - dest.stat().st_mtime
        if age < CACHE_FRESH_SECONDS:
            print(f"  Cache hit (<{CACHE_FRESH_SECONDS // 60} min old) — reusing {dest.name}, no download")
            return dest
        # older than the window: reuse only if the Sheet hasn't changed in Drive
        cached_mod = _load_meta(meta_path).get("sheet_modified")
        current_mod = _sheet_modified_time(sheet_id)
        if cached_mod and current_mod and cached_mod == current_mod:
            os.utime(dest, None)  # restart the freshness window
            print(f"  Sheet unchanged since last fetch — reusing {dest.name}")
            return dest
        print("  Sheet changed (or freshness check failed) — re-downloading...")

    # --- download: team tab + month-tab candidates run concurrently; the
    # modifiedTime describe is folded into the same pool so it costs no extra
    # wall-clock (skipped if we already fetched it during the freshness check) ---
    with ThreadPoolExecutor(max_workers=len(candidates) + 2) as ex:
        f_team = ex.submit(_read_tab, sheet_id, TEAM_TAB, TEAM_RANGE)
        f_month = {c: ex.submit(_try_read_tab, sheet_id, c, MONTH_RANGE) for c in candidates}
        f_mod = ex.submit(_sheet_modified_time, sheet_id) if current_mod is None else None
        team_rows = f_team.result()
        month_title = month_rows = None
        for cand in candidates:
            rows = f_month[cand].result()
            if rows:
                month_title, month_rows = cand, rows
                break
        if f_mod is not None:
            current_mod = f_mod.result()

    if month_rows is None:
        raise RuntimeError(f"No monthly tab found for {month} {year} (tried {candidates}).")

    wb = Workbook()
    wb.remove(wb.active)  # drop the default empty sheet
    _write_tab(wb, TEAM_TAB, team_rows)
    _write_tab(wb, month_title, month_rows)
    dest.parent.mkdir(parents=True, exist_ok=True)
    wb.save(dest)

    if current_mod:
        meta_path.write_text(json.dumps({"sheet_modified": current_mod, "fetched_at": time.time()}))
    return dest
