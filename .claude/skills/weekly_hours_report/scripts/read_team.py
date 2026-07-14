from pathlib import Path
from openpyxl import load_workbook

SHEET = "Team allocation 2026"
HEADER_ROW = 1

# report field -> column header in the "Team allocation 2026" tab
COLS = {
    "team": "App",                         # product
    "module": "Feature",                   # pilar
    "qa_lead": "QA Lead",
    "job_role": "Role",                    # QA 1 / QA 2 / QA 3 -> Q1/Q2/Q3
    "short_name": "QA Analyst",            # match key vs PTO "Collab" + displayed name
    "full_name": "QA Analyst (Full Name)",
    "qa_3": "New QA3",
}

def read_team(team_path=None):
    if not team_path:
        raise ValueError("team_path (local xlsx fetched from the Google Sheet) is required")
    p = Path(team_path)
    if not p.exists(): raise FileNotFoundError(f"Team file not found at {p}")
    wb=load_workbook(p,data_only=True); ws=wb[SHEET]; headers={}
    for col in range(1,ws.max_column+1):
        v=ws.cell(HEADER_ROW,col).value
        if v: headers[str(v).strip()]=col
    out=[]
    for r in range(HEADER_ROW+1,ws.max_row+1):
        def get(h): c=headers.get(h); return str(ws.cell(r,c).value or "").strip() if c else ""
        short=get(COLS["short_name"])
        if not short: continue
        rec={field:get(h) for field,h in COLS.items()}
        # fields absent from the new Sheet, kept empty for downstream compatibility
        rec.update({"wave":"","meta_username":"","meta_tag":"","rate":"","softtek_username":""})
        out.append(rec)
    return out
