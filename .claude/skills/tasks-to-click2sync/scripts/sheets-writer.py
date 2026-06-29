#!/usr/bin/env python3
"""Writes C2CRow data to Click2SyncReport.xlsx.
Strategy: copy from Shared Drive to /tmp, modify with openpyxl, cp back.
Includes retry + verification to handle sync conflicts."""

import json
import os
import re
import shutil
import sys
import time
import tempfile
from datetime import datetime, timedelta
from pathlib import Path

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    print("Error: openpyxl not installed. Run: pip3 install openpyxl", file=sys.stderr)
    sys.exit(1)

SKILL_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
CONFIG_PATH = os.path.join(SKILL_DIR, "config.json")

LOCK_STALE_TIMEOUT = 60
LOCK_MAX_WAIT = 120
LOCK_POLL_INTERVAL = 3
MAX_RETRIES = 3
RETRY_DELAY = 10

COLUMNS = [
    ("projectId", "Project ID", "default"),
    ("createdBy", "Created By", "default"),
    ("requestNo", "Request No", "default"),
    ("assignTo", "Assigned To", "gold"),
    ("peerReviewer", "Peer Reviewer", "default"),
    ("shortDescription", "Short Description", "default"),
    ("requirementType", "Requirement Type", "default"),
    ("requirementSubtype", "Requirement Subtype", "default"),
    ("moduleFeature", "Module/Feature", "gold"),
    ("team", "Team", "gold"),
    ("numberOfDefectiveProducts", "Number of Defective Products", "default"),
    ("numberOfProducts", "Number of Products (Generated or Modified)", "gold"),
    ("totalOfTestCases", "Total of Test Cases", "default"),
    ("total", "Total", "default"),
    ("pass", "Pass", "default"),
    ("fail", "Fail", "default"),
    ("cantTest", "Cant Test", "default"),
    ("qtyBugsClosed", "Quantity of Bugs Closed", "default"),
    ("qtyBugsReopened", "Quantity of Bugs Re-opened", "default"),
    ("qtyBugsVerified", "Quantity of Bugs Verified", "default"),
    ("qtyNewBugsFound", "Quantity of New Bugs Found", "default"),
    ("releaseBuild", "Release/Build", "gold"),
    ("completePercent", "Complete %", "default"),
    ("defectsAdviceFlag", "Defects Advice Flag", "default"),
    ("peerReviewChecklistReviewed1", "Peer Review Checklist Reviewed", "default"),
    ("peerReviewChecklistTemplate", "Peer Review Checklist Template", "default"),
    ("peerReviewChecklistReviewed2", "Peer Review Checklist Reviewed", "default"),
    ("scheduledStartDate", "Scheduled Start Date", "orange"),
    ("scheduledFinishDate", "Scheduled Finish Date", "orange"),
    ("scheduledDuration", "Scheduled Duration", "default"),
    ("scheduledEffort", "Scheduled Effort", "default"),
    ("peerReviewScheduledEffort", "Peer Review Scheduled Effort (hrs)", "green"),
    ("scheduledReworkEffort", "Scheduled Rework Effort (hrs)", "green"),
    ("scheduledUATReworkEffort", "Scheduled UAT Rework Effort", "green"),
    ("actualStartDate", "Actual Start Date", "default"),
    ("actualFinishDate", "Actual Finish Date", "default"),
    ("actualDuration", "Actual Duration", "default"),
    ("actualEffort", "Actual Effort", "default"),
    ("peerReviewActualEffort", "Peer Review Actual Effort (hrs)", "green"),
    ("actualReworkEffort", "Actual Rework Effort (hrs)", "green"),
    ("actualUATReworkEffort", "Actual UAT Rework Effort", "green"),
    ("closedOn", "Closed On", "default"),
    ("forClosing", "ForClosing", "default"),
    ("action", "Action", "default"),
    ("peerReviewChecklist", "Peer Review Checklist Reviewed", "blueMedium"),
    ("estimationLink", "Estimation", "blueMedium"),
]

COLORS = {
    "default": "1a3a5c",
    "gold": "8b6d1a",
    "green": "2d5a3a",
    "blueMedium": "1a4b6e",
    "orange": "8b4513",
}

GRAYABLE_FIELDS = {
    "estimationLink", "peerReviewer", "peerReviewScheduledEffort",
    "peerReviewActualEffort", "peerReviewChecklist", "numberOfProducts",
    "numberOfDefectiveProducts", "totalOfTestCases", "actualReworkEffort",
    "actualUATReworkEffort", "scheduledReworkEffort", "scheduledUATReworkEffort",
    "defectsAdviceFlag", "peerReviewChecklistReviewed1",
    "peerReviewChecklistTemplate", "peerReviewChecklistReviewed2", "closedOn",
    "total", "pass", "fail", "cantTest", "qtyBugsClosed",
    "qtyBugsReopened", "qtyBugsVerified", "qtyNewBugsFound",
}

THIN_BORDER = Border(
    top=Side(style='thin'), bottom=Side(style='thin'),
    left=Side(style='thin'), right=Side(style='thin')
)


def get_week_tab_name():
    today = datetime.now()
    monday = today - timedelta(days=today.weekday())
    sunday = monday + timedelta(days=6)
    return f"{monday.strftime('%b')} {monday.day}-{sunday.day}"


def get_xlsx_path():
    cloud_storage = Path.home() / "Library" / "CloudStorage"
    gdrive_dirs = list(cloud_storage.glob("GoogleDrive-*@meta.com"))
    if not gdrive_dirs:
        print("Error: Google Drive not found in ~/Library/CloudStorage/", file=sys.stderr)
        sys.exit(1)
    base = gdrive_dirs[0]
    xlsx_path = base / "Shared drives" / "Meta - STK" / "Project Tracking" / "Automation" / "Automation Outputs" / "Click2SyncReport.xlsx"
    if not xlsx_path.exists():
        print(f"Error: File not found: {xlsx_path}", file=sys.stderr)
        sys.exit(1)
    return xlsx_path


def acquire_lock(lock_path, username):
    print("Acquiring write lock...", end=" ", flush=True)
    waited = 0
    while lock_path.exists():
        try:
            lock_content = lock_path.read_text().strip()
            lock_time = lock_content.split("|")[-1] if "|" in lock_content else ""
            if lock_time:
                lock_dt = datetime.fromisoformat(lock_time)
                age = (datetime.now() - lock_dt).total_seconds()
                if age > LOCK_STALE_TIMEOUT:
                    lock_path.unlink(missing_ok=True)
                    break
        except (ValueError, OSError):
            pass
        if waited >= LOCK_MAX_WAIT:
            print("TIMEOUT")
            print("Error: Could not acquire lock. Try again in 2 minutes.", file=sys.stderr)
            sys.exit(1)
        time.sleep(LOCK_POLL_INTERVAL)
        waited += LOCK_POLL_INTERVAL

    lock_path.write_text(f"{username}|{datetime.now().isoformat()}")
    time.sleep(2)
    try:
        content = lock_path.read_text().strip()
        if not content.startswith(f"{username}|"):
            print("FAILED")
            print("Error: Lock taken by another user.", file=sys.stderr)
            sys.exit(1)
    except OSError:
        pass
    print("OK")


def release_lock(lock_path):
    lock_path.unlink(missing_ok=True)


def get_last_data_row(ws):
    last = 1
    for row_idx in range(2, ws.max_row + 1):
        if ws.cell(row=row_idx, column=2).value:
            last = row_idx
    return last


def get_last_request_no(wb, tab_name):
    if tab_name in wb.sheetnames:
        ws = wb[tab_name]
        for row_idx in range(ws.max_row, 0, -1):
            val = ws.cell(row=row_idx, column=3).value
            if val and str(val).startswith("29582-3-"):
                try:
                    return int(str(val).replace("29582-3-", ""))
                except ValueError:
                    continue
    for tab in reversed(wb.sheetnames):
        if tab == tab_name:
            continue
        ws = wb[tab]
        for row_idx in range(ws.max_row, 0, -1):
            val = ws.cell(row=row_idx, column=3).value
            if val and str(val).startswith("29582-3-"):
                try:
                    return int(str(val).replace("29582-3-", ""))
                except ValueError:
                    continue
    return 10200


def extract_task_id(short_description):
    match = re.match(r'^\[T?(\d+)\]\s*', str(short_description))
    return match.group(1) if match else None


def write_headers(ws):
    for col_idx, (key, label, color_name) in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.font = Font(bold=True, color="FFFFFF", size=9)
        cell.fill = PatternFill(start_color=COLORS[color_name], end_color=COLORS[color_name], fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER
    ws.row_dimensions[1].height = 30
    last_col_letter = ws.cell(row=1, column=len(COLUMNS)).column_letter
    ws.auto_filter.ref = f"A1:{last_col_letter}1"


def write_row(ws, row_idx, row_data):
    for col_idx, (key, _, _) in enumerate(COLUMNS, 1):
        value = row_data.get(key, "")
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.font = Font(size=9)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER
        if key in GRAYABLE_FIELDS and not value:
            cell.fill = PatternFill(start_color="d4d4d8", end_color="d4d4d8", fill_type="solid")
            cell.font = Font(size=9, color="71717a")


def set_column_widths(ws):
    for col_idx, (key, label, _) in enumerate(COLUMNS, 1):
        col_letter = ws.cell(row=1, column=col_idx).column_letter
        ws.column_dimensions[col_letter].width = len(label) + 4


def row_has_changes(ws, row_idx, row_data):
    for col_idx, (key, _, _) in enumerate(COLUMNS, 1):
        if key == "requestNo":
            continue
        old_val = str(ws.cell(row=row_idx, column=col_idx).value or "")
        new_val = str(row_data.get(key, ""))
        if old_val != new_val:
            return True
    return False


def verify_write(xlsx_path, tab_name, my_username, expected_task_ids):
    """Re-read the file from Shared Drive and check our rows are there."""
    try:
        wb = load_workbook(str(xlsx_path), read_only=True)
        if tab_name not in wb.sheetnames:
            wb.close()
            return False
        ws = wb[tab_name]
        found_ids = set()
        for row_idx in range(2, ws.max_row + 1):
            owner = ws.cell(row=row_idx, column=2).value
            desc = ws.cell(row=row_idx, column=6).value
            if owner and str(owner) == my_username and desc:
                task_id = extract_task_id(desc)
                if task_id:
                    found_ids.add(task_id)
        wb.close()
        return expected_task_ids.issubset(found_ids)
    except Exception:
        return False


def do_write(rows, my_username, tab_name, xlsx_path, tmp_path):
    """Copy, modify, copy back. Returns (success, message)."""
    # Copy from Shared Drive to /tmp
    shutil.copy2(str(xlsx_path), str(tmp_path))

    # Open and modify
    wb = load_workbook(str(tmp_path))

    if tab_name not in wb.sheetnames:
        ws = wb.create_sheet(tab_name)
        write_headers(ws)
    else:
        ws = wb[tab_name]

    last_req_no = get_last_request_no(wb, tab_name)
    counter = last_req_no + 1

    existing_task_rows = {}
    for row_idx in range(2, ws.max_row + 1):
        owner = ws.cell(row=row_idx, column=2).value
        desc = ws.cell(row=row_idx, column=6).value
        if owner and str(owner) == my_username and desc:
            task_id = extract_task_id(desc)
            if task_id:
                existing_task_rows[task_id] = row_idx

    updated_count = 0
    new_rows = []

    for row in rows:
        task_id = extract_task_id(row["shortDescription"])
        if task_id and task_id in existing_task_rows:
            target_row_idx = existing_task_rows[task_id]
            if row_has_changes(ws, target_row_idx, row):
                existing_req_no = ws.cell(row=target_row_idx, column=3).value
                if existing_req_no:
                    row["requestNo"] = str(existing_req_no)
                write_row(ws, target_row_idx, row)
                updated_count += 1
        else:
            row["requestNo"] = f"29582-3-{counter:05d}"
            new_rows.append(row)
            counter += 1

    if new_rows:
        start_row = get_last_data_row(ws) + 1
        for i, row_data in enumerate(new_rows):
            write_row(ws, start_row + i, row_data)

    set_column_widths(ws)

    if updated_count == 0 and not new_rows:
        return True, f"All rows up to date for '{my_username}' in '{tab_name}'."

    # Save and copy back
    wb.save(str(tmp_path))
    shutil.copy2(str(tmp_path), str(xlsx_path))

    parts = []
    if new_rows:
        parts.append(f"{len(new_rows)} new")
    if updated_count:
        parts.append(f"{updated_count} updated")
    return True, f"{', '.join(parts)} rows in '{tab_name}' in Click2SyncReport.xlsx"


def main():
    if not os.path.exists(CONFIG_PATH):
        print("Error: config.json not found.", file=sys.stderr)
        sys.exit(1)

    with open(CONFIG_PATH) as f:
        config = json.load(f)

    rows = json.loads(sys.stdin.read())
    if not rows:
        print("No rows to write.")
        return

    xlsx_path = get_xlsx_path()
    tab_name = get_week_tab_name()
    my_username = config["softtek_username"]
    lock_path = xlsx_path.parent / "c2c_write.lock"
    tmp_path = Path(tempfile.gettempdir()) / "Click2SyncReport_edit.xlsx"

    # Collect expected task IDs for verification
    expected_task_ids = set()
    for row in rows:
        tid = extract_task_id(row.get("shortDescription", ""))
        if tid:
            expected_task_ids.add(tid)

    # Acquire lock
    acquire_lock(lock_path, my_username)

    try:
        for attempt in range(1, MAX_RETRIES + 1):
            if attempt > 1:
                print(f"Retry {attempt}/{MAX_RETRIES} (waiting {RETRY_DELAY}s)...", flush=True)
                time.sleep(RETRY_DELAY)

            print("Writing rows...", flush=True)
            success, message = do_write(rows, my_username, tab_name, xlsx_path, tmp_path)

            if not success:
                print(f"Attempt {attempt} failed: {message}", file=sys.stderr)
                continue

            if "up to date" in message:
                print(message)
                tmp_path.unlink(missing_ok=True)
                return

            # Verify: wait a moment then re-read the file
            time.sleep(3)
            if verify_write(xlsx_path, tab_name, my_username, expected_task_ids):
                print(message)
                tmp_path.unlink(missing_ok=True)
                return
            else:
                print(f"Attempt {attempt}: Verification failed, retrying...", file=sys.stderr)

        # All retries failed
        print("=" * 60, file=sys.stderr)
        print("  WRITE FAILED after all retries.", file=sys.stderr)
        print("  Your rows were NOT saved. Wait 1 minute and try again.", file=sys.stderr)
        print("=" * 60, file=sys.stderr)
        tmp_path.unlink(missing_ok=True)
        sys.exit(1)

    finally:
        release_lock(lock_path)


if __name__ == "__main__":
    main()
